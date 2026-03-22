import face_recognition
import cv2
import numpy as np
import csv
import io
import os
import threading
from datetime import datetime
from flask import Flask, render_template, Response, jsonify, send_file, request, redirect, url_for, session, flash
from openpyxl import Workbook
from functools import wraps

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here-change-in-production'

# Subjects
SUBJECTS = ['WPPM', 'MPV', 'VEP', 'TECHNICAL', 'ENGLISH']

# ---------------------------------------------------------------------------
# Global state
# ---------------------------------------------------------------------------
PHOTOS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                          "LEARNING PROJECT", "PHOTOS")
ATTENDANCE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              "attendance_records")

known_face_encodings = []
known_face_names = []

# Attendance for the current session  {name: {subject: {'status': 'present/absent', 'time': time_string}}}
attendance_record: dict[str, dict[str, dict[str, str]]] = {}
total_students: list[str] = []

camera_active = False
video_capture = None
lock = threading.Lock()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _ensure_dirs():
    os.makedirs(ATTENDANCE_DIR, exist_ok=True)


def load_known_faces():
    """Load every .jpg/.png in PHOTOS_DIR and encode the faces."""
    global known_face_encodings, known_face_names, total_students

    known_face_encodings = []
    known_face_names = []

    if not os.path.isdir(PHOTOS_DIR):
        raise FileNotFoundError(f"Photos directory not found: {PHOTOS_DIR}")

    for fname in sorted(os.listdir(PHOTOS_DIR)):
        if not fname.lower().endswith((".jpg", ".jpeg", ".png")):
            continue
        path = os.path.join(PHOTOS_DIR, fname)
        image = face_recognition.load_image_file(path)
        # Normalize to 8-bit RGB (handles RGBA PNGs, grayscale, numpy 2.x dtype issues)
        if len(image.shape) == 2:
            image = np.stack([image] * 3, axis=-1)
        elif image.shape[2] == 4:
            image = image[:, :, :3]
        image = np.ascontiguousarray(image, dtype=np.uint8)
        encodings = face_recognition.face_encodings(image)
        if encodings:
            known_face_encodings.append(encodings[0])
            name = os.path.splitext(fname)[0].upper()
            known_face_names.append(name)

    if not known_face_names:
        raise RuntimeError("No faces found in any photo.")

    total_students = known_face_names.copy()


def _initialize_attendance():
    """Initialize attendance record for all students and subjects."""
    global attendance_record
    attendance_record = {}
    for student in total_students:
        attendance_record[student] = {}
        for subject in SUBJECTS:
            attendance_record[student][subject] = {'status': 'absent', 'time': ''}


def _save_excel():
    """Persist today's attendance to an Excel file."""
    _ensure_dirs()
    date_str = datetime.now().strftime("%Y-%m-%d")
    path = os.path.join(ATTENDANCE_DIR, f"{date_str}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Headers
    ws['A1'] = 'Serial No.'
    ws['B1'] = 'Name'
    ws['C1'] = 'WPPM'
    ws['D1'] = 'MPV'
    ws['E1'] = 'VEP'
    ws['F1'] = 'TECHNICAL'
    ws['G1'] = 'ENGLISH'

    row = 2
    for idx, student in enumerate(total_students, 1):
        ws[f'A{row}'] = idx
        ws[f'B{row}'] = student
        ws[f'C{row}'] = attendance_record[student]['WPPM']['status'].title()
        ws[f'D{row}'] = attendance_record[student]['MPV']['status'].title()
        ws[f'E{row}'] = attendance_record[student]['VEP']['status'].title()
        ws[f'F{row}'] = attendance_record[student]['TECHNICAL']['status'].title()
        ws[f'G{row}'] = attendance_record[student]['ENGLISH']['status'].title()
        row += 1

    wb.save(path)


def generate_frames():
    """Yield MJPEG frames with face-recognition overlay."""
    global camera_active, video_capture

    while camera_active:
        if video_capture is None or not video_capture.isOpened():
            break

        ret, frame = video_capture.read()
        if not ret:
            break

        # Resize for faster recognition
        small = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_small = np.ascontiguousarray(
            cv2.cvtColor(small, cv2.COLOR_BGR2RGB), dtype=np.uint8)

        face_locations = face_recognition.face_locations(rgb_small)
        face_names = []

        if face_locations:
            face_encodings = face_recognition.face_encodings(rgb_small,
                                                             face_locations)
            for encoding in face_encodings:
                matches = face_recognition.compare_faces(
                    known_face_encodings, encoding)
                name = "Unknown"

                if known_face_encodings:
                    distances = face_recognition.face_distance(
                        known_face_encodings, encoding)
                    best_idx = np.argmin(distances)
                    if matches[best_idx]:
                        name = known_face_names[best_idx]

                face_names.append(name)

                # Record attendance for all subjects
                with lock:
                    if name != "Unknown" and name in attendance_record:
                        time_str = datetime.now().strftime("%H:%M:%S")
                        for subject in SUBJECTS:
                            if attendance_record[name][subject]['status'] == 'absent':
                                attendance_record[name][subject] = {'status': 'present', 'time': time_str}
                        _save_excel()

        # Draw boxes
        for (top, right, bottom, left), name in zip(face_locations,
                                                    face_names):
            top *= 4
            right *= 4
            bottom *= 4
            left *= 4
            color = (0, 255, 0) if name != "Unknown" else (0, 0, 255)
            cv2.rectangle(frame, (left, top), (right, bottom), color, 2)
            cv2.rectangle(frame, (left, bottom - 35), (right, bottom),
                          color, cv2.FILLED)
            cv2.putText(frame, name, (left + 6, bottom - 6),
                        cv2.FONT_HERSHEY_DUPLEX, 0.8, (255, 255, 255), 1)

        _, buffer = cv2.imencode('.jpg', frame)
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' +
               buffer.tobytes() + b'\r\n')


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    if 'logged_in' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route("/login", methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        # Simple authentication - in production, use proper authentication
        if username == 'admin' and password == 'Admin123':
            session['logged_in'] = True
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid credentials')
    return render_template('login.html')


@app.route("/logout")
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('login'))


@app.route("/dashboard")
@login_required
def dashboard():
    return render_template("index.html")


@app.route("/video_feed")
@login_required
def video_feed():
    if not camera_active:
        return "", 204
    return Response(generate_frames(),
                    mimetype="multipart/x-mixed-replace; boundary=frame")


@app.route("/start", methods=["POST"])
@login_required
def start():
    global camera_active, video_capture, attendance_record

    if camera_active:
        return jsonify(status="already_running")

    try:
        load_known_faces()
        _initialize_attendance()
    except Exception as e:
        return jsonify(status="error", message=str(e)), 500

    video_capture = cv2.VideoCapture(0)
    if not video_capture.isOpened():
        return jsonify(status="error",
                       message="Could not open camera."), 500

    camera_active = True
    return jsonify(status="started",
                   total=len(total_students),
                   students=total_students)


@app.route("/stop", methods=["POST"])
@login_required
def stop():
    global camera_active, video_capture

    camera_active = False
    if video_capture is not None:
        video_capture.release()
        video_capture = None

    _save_excel()
    return jsonify(status="stopped")


@app.route("/attendance")
@login_required
def attendance():
    """Return current attendance data as JSON."""
    with lock:
        present_count = 0
        absent_count = 0
        attendance_data = []

        for name in total_students:
            student_present = False
            for subject in SUBJECTS:
                if attendance_record[name][subject]['status'] == 'present':
                    student_present = True
                    break
            if student_present:
                present_count += 1
            else:
                absent_count += 1

            attendance_data.append({
                "name": name,
                "subjects": attendance_record[name]
            })

    return jsonify(
        total=len(total_students),
        present_count=present_count,
        absent_count=absent_count,
        attendance=attendance_data,
        date=datetime.now().strftime("%Y-%m-%d"),
    )


@app.route("/download_excel")
@login_required
def download_excel():
    """Generate and send attendance Excel for download."""
    with lock:
        mem = io.BytesIO()

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"

        # Headers
        ws['A1'] = 'Serial No.'
        ws['B1'] = 'Name'
        ws['C1'] = 'WPPM'
        ws['D1'] = 'MPV'
        ws['E1'] = 'VEP'
        ws['F1'] = 'TECHNICAL'
        ws['G1'] = 'ENGLISH'

        row = 2
        date_str = datetime.now().strftime("%Y-%m-%d")
        for idx, student in enumerate(total_students, 1):
            ws[f'A{row}'] = idx
            ws[f'B{row}'] = student
            ws[f'C{row}'] = attendance_record[student]['WPPM']['status'].title()
            ws[f'D{row}'] = attendance_record[student]['MPV']['status'].title()
            ws[f'E{row}'] = attendance_record[student]['VEP']['status'].title()
            ws[f'F{row}'] = attendance_record[student]['TECHNICAL']['status'].title()
            ws[f'G{row}'] = attendance_record[student]['ENGLISH']['status'].title()
            row += 1

        wb.save(mem)
        mem.seek(0)

        filename = f"attendance_{date_str}.xlsx"
        return send_file(mem, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True,
                         download_name=filename)


# ---------------------------------------------------------------------------
if __name__ == "__main__":
    _ensure_dirs()
    port = int(os.environ.get("PORT", 5001))
    print("Starting Smart Attendance System...")
    print(f"Open http://127.0.0.1:{port} in your browser")
    app.run(host="0.0.0.0", port=port, debug=False, threaded=True)
